import pandas as pd
from nltk.tokenize import word_tokenize  # 分词器加载
import re
from nltk.corpus import stopwords
import openpyxl



github_standard_data = pd.read_excel("github_gold.xlsx", header=None)
test_data = pd.DataFrame(github_standard_data)
# print(test_data[2])


'''
数据清洗  data clean
'''
english_stopwords = stopwords.words('english')
def text_clean(text):
    # print("原始数据：", text, '\n')

    #去掉HTML标签（e.g. &temp;）
    text_no_special_entities = re.sub(r'\&\w*;|#\w*|@\w*', '', text)
    # print("去掉标签后：", text_no_special_entities, '\n')

    #去掉价值符号
    text_no_tickers = re.sub(r'\$\w*', '', text_no_special_entities)
    # print("去掉价值符号后：", text_no_tickers, '\n')

    #去掉超链接
    text_no_hyperlinks = re.sub(r'https?:\/\/.*\/\w*', '', text_no_tickers)
    # print("去掉超链接后：", text_no_hyperlinks, '\n')

    #去掉一些专门的名词缩写，即字母较少的词
    text_no_small_words = re.sub(r'\b\w{1,2}\b', '', text_no_hyperlinks)
    # print("去掉专有名词后：", text_no_small_words, '\n')

    #去掉多余的空格
    text_no_whitespace = re.sub(r'\s\s+', ' ', text_no_small_words)
    text_no_whitespace = text_no_whitespace.lstrip(' ')
    # print("去掉多余空格后：", text_no_whitespace, '\n')

    #分词
    tokens = word_tokenize(text_no_whitespace)
    # print("分词结果：", tokens, '\n')

    #去除标点符号
    filtered_isalpha = [word.lower() for word in tokens if word.isalpha()]
    # print("去除标点符号后：", filtered_isalpha, '\n')

    #去除停用词
    list_no_stopwords = [i for i in filtered_isalpha if i not in english_stopwords]
    print("去除停用词后：", list_no_stopwords)
    #过滤后结果
    text_filtered = ' '.join(list_no_stopwords)
    print("过滤后：", text_filtered)
    return text_filtered

wb = openpyxl.load_workbook('github_gold.xlsx')
sheet = wb.worksheets[0]
sheet.cell(2,4,"Standard_Comment_Data_cleansing")
for i in range(2,len(test_data[2])):
    data_proceed = text_clean(test_data[2][i])
    sheet.cell(i+1, 4, data_proceed)
    print("正在写入第"+str(i)+"条")


'''
增加列： pos = 1， neg = -1, neu = 0
'''

sheet.cell(2,5,"Polarity_number")
for i in range(2,len(test_data[2])):
    polarity_data_read = test_data[1][i]
    if polarity_data_read == 'neutral':
        sheet.cell(i + 1, 5, 0)
        print(polarity_data_read)
        print("正在写入第" + str(i) + "条")
    elif polarity_data_read == 'positive':
        sheet.cell(i + 1, 5, 1)
        print(polarity_data_read)
        print("正在写入第" + str(i) + "条")
    else:
        sheet.cell(i + 1, 5, -1)
        print(polarity_data_read)
        print("正在写入第" + str(i) + "条")
    # print(polarity_data_read)
    # sheet.cell(i+1, 4, polarity_data_read)
    # print("正在写入第"+str(i)+"条")


wb.save("github_standard_clean.xlsx")