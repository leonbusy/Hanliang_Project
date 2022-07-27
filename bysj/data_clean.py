import pandas as pd
import re
import nltk
from nltk.tokenize import word_tokenize  # 分词器加载
from nltk.text import Text
from nltk.corpus import stopwords
from nltk import pos_tag # 词性标注库
from nltk import ne_chunk
from sklearn.feature_extraction.text import CountVectorizer
from sklearn.model_selection import train_test_split
from sklearn.svm import SVC
import openpyxl
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
import time

# nltk.download('words') # pip后安装了基本股价 nltk.download()选择性安装更多包 : words, punkt, stopwords, averaged_perceptron_tagger, maxent_ne_chunker

df = pd.read_excel("data_final.xlsx",sheet_name="sheet1", header=None)
# sentence = df[4][2]


'''
句子分词
'''
# tokens = word_tokenize(sentence)  # 句子分词
# print(tokens[:10])

'''
对分词创建Text对象,可以对文本操作
'''
# t = Text(tokens)
# t.plot(10) #图形展示前十个词词频的分布

'''
过滤停用词
'''
# print(stopwords.raw('english').replace('\n', ' ')) # 停用词词库的词有哪些
# test_words_set = set(tokens)
# print(test_words_set.intersection(set(stopwords.words('english'))))  # 查看分词后与停用词表有哪些交集的词
# filtered = [w for w in test_words_set if (w not in stopwords.words('english'))] # 过滤停用词后剩余词
# filtered_isalpha = [word.lower() for word in filtered if word.isalpha()]  # 过滤标点符号和特殊符号
# print(tokens)  # 打印原始分词
# print(filtered) # 打印过滤停用词
# print(filtered_isalpha) # 打印过滤字符

'''
词性标注
'''
# words_tags = pos_tag(filtered_isalpha)
# print(words_tags)

'''
命名实体识别
'''
# words_ne_chunk = ne_chunk(words_tags)
# print(words_ne_chunk)




'''
数据清洗  data clean
'''
english_stopwords = stopwords.words('english')
def text_clean(text):
    # print("原始数据：", text, '\n')
    # 去掉标签 如<strong></strong>
    text_no_entities = re.sub(r'\<(.*?)docs\</a\>.|\<code\>(.*?)\</code\>|\<img(.*?)\>|\<strong\>(.*?)\</strong\>|\<em\>(.*?)\</em\>','',str(text))
    print(text_no_entities)

    #去掉HTML标签（e.g. &temp;）
    text_no_special_entities = re.sub(r'\&\w*;|#\w*|@\w*', '', text_no_entities)
    print("去掉标签后：", text_no_special_entities, '\n')

    #去掉价值符号
    text_no_tickers = re.sub(r'\$\w*', '', text_no_special_entities)
    # print("去掉价值符号后：", text_no_tickers, '\n')

    #去掉超链接
    text_no_hyperlinks = re.sub(r'https?:\/\/.*\/\w*', '', text_no_tickers)
    # print("去掉超链接后：", text_no_hyperlinks, '\n')

    #去掉一些专门的名词缩写，即字母较少的词
    text_no_small_words = re.sub(r'\b\w{1,2}\b', '', text_no_hyperlinks)
    # print("去掉专有名词后：", text_no_small_words, '\n')

    #去掉多余的空格
    text_no_whitespace = re.sub(r'\s\s+', ' ', text_no_small_words)
    text_no_whitespace = text_no_whitespace.lstrip(' ')
    # print("去掉多余空格后：", text_no_whitespace, '\n')

    #分词
    tokens = word_tokenize(text_no_whitespace)
    # print("分词结果：", tokens, '\n')

    #去除标点符号
    filtered_isalpha = [word.lower() for word in tokens if word.isalpha()]
    # print("去除标点符号后：", filtered_isalpha, '\n')

    #去除停用词
    list_no_stopwords = [i for i in filtered_isalpha if i not in english_stopwords]
    print("去除停用词后：", list_no_stopwords)
    #过滤后结果
    text_filtered = ' '.join(list_no_stopwords)
    print("过滤后：", text_filtered)
    return text_filtered

wb = openpyxl.load_workbook('data_final.xlsx')
sheet = wb.worksheets[0]
sheet.cell(1,9,"Comment_Data_cleansing")
for i in range(1,len(df[5])):
    data_proceed = text_clean(df[5][i])
    sheet.cell(i + 1, 9, str(data_proceed))
    print("正在写入第" + str(i) + "条")

wb.save("data_final_clean2.xlsx")


