from sklearn.externals import joblib
import pandas as pd
import openpyxl


get_data = pd.read_excel("data_final_final_clean.xlsx", header=None,keep_default_na=False)

wb = openpyxl.load_workbook('data_final_final_clean.xlsx')
sheet = wb.worksheets[0]
sheet.cell(1, 10, "Comment_predict_number")
sheet.cell(1, 11, "Comment_predict_sentiment")


for i in range(1,len(get_data[8])):
    data_read = get_data[8][i]
    # print(data_read)
    if data_read == '':
        data_read = "None"
        # print(data_read)
    model = joblib.load('tfidf_svm_sentiment.model')
    y_pre = model.predict([data_read])
    print(y_pre)
    proba = model.predict_proba([data_read])[0]
    print(proba)

    if y_pre[0] == -1:
        print(data_read,": Most likely negative in sentiment（Probability："+str(proba[0])+")")
        sheet.cell(i+1,10, int(y_pre[0]))
        sheet.cell(i+1,11,"negative")
    elif y_pre[0] == 0:
        print(data_read,": Most likely neutral in sentiment（Probability："+str(proba[1])+")")
        sheet.cell(i+1,10, int(y_pre[0]))
        sheet.cell(i+1,11,"neutral")
    elif y_pre[0] == 1:
        print(data_read,": Most likely positive in sentiment（Probability："+str(proba[2])+")")
        sheet.cell(i+1,10, int(y_pre[0]))
        sheet.cell(i+1,11,"positive")



# for i in range(1,len(df[4])):
#     data_proceed = text_clean(df[4][i])
#     sheet.cell(i+1,7,data_proceed)
#     print("正在写入第"+str(i)+"条")

wb.save("data_final_clean_predict.xlsx")